import re
from convertdate import indian_civil
from convertdate import islamic
from convertdate import julian
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

date_column = 1  # Specify the column number where the dates are located
start_row = 2  # Specify the starting row for dates (e.g., 2 for skipping the header row)

wb = load_workbook('converted_one_2.xlsx')
ws = wb.active
for row in range(start_row, ws.max_row + 1):
    date_string = ws[get_column_letter(date_column) + str(row)].value

    if date_string is None:
        continue

    if isinstance(date_string, datetime):
        date_string = date_string.strftime("%d-%m-%y")

    if re.match(r'\d{1,2}/\d{1,2}/\d{4}', date_string):
        dt = datetime.strptime(date_string, "%m/%d/%Y")
    elif re.match(r'\d{1,2}-\d{1,2}-\d{2}', date_string):
        dt = datetime.strptime(date_string, "%d-%m-%y")
    else:
        continue

    day = dt.day
    month = dt.month
    year = dt.year

    indian_date = indian_civil.from_gregorian(year, month, day)
    indian_date_string = str(indian_date[0]) + "-" + str(indian_date[1]) + "-" + str(indian_date[2])
    islamic_date = islamic.from_gregorian(year,month,day)
    islamic_date_string = str(islamic_date[0]) + "-" + str(islamic_date[1])+ "-" + str(islamic_date[2])
    julian_date = julian.from_gregorian(year,month,day)
    julian_date_string = str(julian_date[0]) + "-" + str(julian_date[1])+ "-" + str(julian_date[2])
    ws[get_column_letter(6) + str(row)] = indian_date_string
    ws[get_column_letter(8)+ str(row)]  = islamic_date_string
    ws[get_column_letter(10)+ str(row)]  = julian_date_string
    

wb.save('converted_one_2.xlsx')
